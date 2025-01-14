import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import glob
import locale

# Configurar locale para formatação de números
locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')

def read_and_process_file(filepath):
    # Nomes das colunas
    colunas = [
        "referencia_externa", "entidade_registradora", "instituicao_credenciadora",
        "usuario_final_recebedor", "arranjo_pagamento", "data_liquidacao",
        "titular_unidade_recebivel", "constituicao_unidade_recebivel", 
        "valor_constituido_total", "valor_constituido_antecipacao_pre_contratado", 
        "valor_bloqueado", "informacoes_pagamento", "carteira", 
        "valor_livre", "valor_total_ur", "dt_atualizacao_ur"
    ]

    # Ler o arquivo usando o delimitador correto
    df = pd.read_csv(filepath, sep=',', header=None, encoding='utf-8')
    
    # Verificar se temos colunas suficientes
    if len(df.columns) >= len(colunas):
        df = df.iloc[:, :len(colunas)]
        df.columns = colunas
    
    return df

def process_payment_data(df):
    # Definir as colunas para informações de pagamento
    colunas_pagamento = [
        "numero_documento_titular", "tipo_conta", "compe", "ispb", 
        "agencia", "numero_conta", "valor_a_pagar", "beneficiario", 
        "data_liquidacao_efetiva", "valor_liquidacao_efetiva", "regra_divisao",
        "valor_onerado_unidade_recebivel", "tipo_informacao_pagamento", 
        "indicador_ordem_efeito", "valor_constituido_contrato_unidade_recebivel"
    ]

    # Processar a coluna informacoes_pagamento
    df['informacoes_pagamento'] = df['informacoes_pagamento'].astype(str)
    info_pagamento = df['informacoes_pagamento'].str.split('|').str[0]
    df_pagamento = info_pagamento.str.split(';', expand=True)
    
    if df_pagamento is not None and len(df_pagamento.columns) >= len(colunas_pagamento):
        df_pagamento = df_pagamento.iloc[:, :len(colunas_pagamento)]
        df_pagamento.columns = colunas_pagamento
    
    # Combinar os DataFrames
    df_final = pd.concat([df.drop('informacoes_pagamento', axis=1), df_pagamento], axis=1)
    
    # Converter colunas de valor para numérico
    colunas_valor = [
        'valor_constituido_total', 'valor_constituido_antecipacao_pre_contratado',
        'valor_bloqueado', 'valor_livre', 'valor_total_ur', 'valor_a_pagar',
        'valor_liquidacao_efetiva', 'valor_onerado_unidade_recebivel',
        'valor_constituido_contrato_unidade_recebivel'
    ]

    for coluna in colunas_valor:
        if coluna in df_final.columns:
            df_final[coluna] = pd.to_numeric(
                df_final[coluna].astype(str)
                .str.replace(',', '.')
                .str.replace('R$', '')
                .str.strip(),
                errors='coerce'
            ).fillna(0)

    # Calcular status de pagamento
    df_final['valor_esperado_por_plano'] = df_final.groupby('usuario_final_recebedor')['valor_a_pagar'].transform('sum') / 2
    df_final['pago'] = df_final['data_liquidacao_efetiva'].notna() & (df_final['data_liquidacao_efetiva'] != '')
    df_final['valor_pago'] = df_final['valor_liquidacao_efetiva']
    df_final['percentual_pago'] = (df_final['valor_pago'] / df_final['valor_esperado_por_plano'] * 100)
    df_final['status_pagamento'] = df_final['percentual_pago'].apply(lambda x: 'PAGO' if x >= 50 else 'NÃO PAGO')

    return df_final

def export_to_excel(df, filename, sheet_name='Sheet1'):
    writer = pd.ExcelWriter(filename, engine='openpyxl')
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    workbook = writer.book
    worksheet = writer.sheets[sheet_name]
    
    # Definir cores para status de pagamento
    cor_pago = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # Verde claro
    cor_nao_pago = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')  # Vermelho claro
    
    # Estilos
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Formatar cabeçalho
    for col in range(1, df.shape[1] + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    # Formatar células de dados
    status_col = df.columns.get_loc('status_pagamento') + 1 if 'status_pagamento' in df.columns else None
    
    for row in range(2, df.shape[0] + 2):
        for col in range(1, df.shape[1] + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.border = border
            
            # Colorir linha baseado no status de pagamento
            if status_col:
                status = worksheet.cell(row=row, column=status_col).value
                fill_color = cor_pago if status == 'PAGO' else cor_nao_pago
                cell.fill = fill_color
            
            valor = df.iloc[row-2, col-1]
            coluna_nome = df.columns[col-1].lower()
            
            if 'valor' in coluna_nome:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right', vertical='center')
            elif 'percentual' in coluna_nome:
                cell.number_format = '#,##0.00"%"'
                cell.alignment = Alignment(horizontal='right', vertical='center')
            elif isinstance(valor, (int, float)):
                cell.alignment = Alignment(horizontal='right', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Ajustar largura das colunas
    for col in range(df.shape[1]):
        max_length = max(
            df[df.columns[col]].astype(str).apply(len).max(),
            len(str(df.columns[col]))
        )
        worksheet.column_dimensions[get_column_letter(col + 1)].width = min(max_length + 2, 50)
    
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.freeze_panes = 'A2'
    
    writer.close()

def main():
    # Caminho do arquivo
    path = 'arquivos_retorno/CERC-AP005_52541797_20241230_0000001_ret'
    arquivos = glob.glob(path)
    
    if not arquivos:
        print("Nenhum arquivo encontrado")
        return
    
    # Ler e processar o arquivo
    df_inicial = read_and_process_file(arquivos[0])
    if df_inicial is None:
        return
    
    # Processar dados de pagamento
    df_processado = process_payment_data(df_inicial)
    
    # Criar resumo por CNPJ
    resumo_cnpj = df_processado.groupby('usuario_final_recebedor').agg({
        'valor_a_pagar': 'sum',
        'valor_pago': 'sum',
        'pago': 'sum',
        'status_pagamento': lambda x: x.value_counts().index[0]
    }).reset_index()
    
    # Calcular percentual total pago
    resumo_cnpj['percentual_total_pago'] = (resumo_cnpj['valor_pago'] / resumo_cnpj['valor_a_pagar'] * 100).round(2)
    
    # Exportar resultados
    export_to_excel(df_processado, 'resultado_analise_pagamentos.xlsx', 'Análise Pagamentos')
    export_to_excel(resumo_cnpj, 'resumo_cnpj_pagamentos.xlsx', 'Resumo CNPJ')

if __name__ == "__main__":
    main()